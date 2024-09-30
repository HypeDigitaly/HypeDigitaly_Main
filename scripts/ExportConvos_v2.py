import requests
import os
import json
from datetime import datetime
import openpyxl  # Přidejte tento import na začátek souboru
from openpyxl.styles import Font, Alignment
import re
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import PatternFill
import random
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image
import numpy as np
import matplotlib.colors as mcolors

# Funkce pro načtení konfigurace ze souboru
def load_config():
    config_file = 'ExportConvos_config.txt'
    config = {}
    with open(config_file, 'r') as config_file:
        for line in config_file:
            key, value = line.strip().split('=', 1)
            if key == 'CATEGORIES':
                config[key] = value.strip('[]').split(',')
            else:
                config[key] = value.strip()
    return config

# Načtení konfigurace
config = load_config()

# Použití načtených hodnot
AUTH_TOKEN = config['AUTH_TOKEN']
PROJECT_ID = config['PROJECT_ID']
START_DATE = config['START_DATE']
END_DATE = config['END_DATE']
OUTPUT_DIRECTORY = config['OUTPUT_DIRECTORY']
CATEGORIES = config['CATEGORIES']

# Constants
BASE_URL = "https://api.voiceflow.com/v2/transcripts"
HEADERS = {
    "Authorization": AUTH_TOKEN,
    "accept": "application/json"
}

def create_output_directory():
    if not os.path.exists(OUTPUT_DIRECTORY):
        os.makedirs(OUTPUT_DIRECTORY)
        print(f"Created directory: {OUTPUT_DIRECTORY}")
    else:
        print(f"Directory already exists: {OUTPUT_DIRECTORY}")

def get_transcript_ids():
    url = f"{BASE_URL}/{PROJECT_ID}"
    params = {
        "startDate": START_DATE,
        "endDate": END_DATE
    }
    response = requests.get(url, headers=HEADERS, params=params)
    response.raise_for_status()
    return [transcript["_id"] for transcript in response.json()]

def get_transcript_dialog(transcript_id):
    url = f"{BASE_URL}/{PROJECT_ID}/{transcript_id}"
    response = requests.get(url, headers=HEADERS)
    response.raise_for_status()
    return response.json()

def extract_messages(dialog):
    messages = []
    for turn in dialog:
        if turn["type"] == "debug" and "payload" in turn and "payload" in turn["payload"]:
            debug_message = turn["payload"]["payload"].get("message", "")
            if "CategoryFilter" in debug_message:
                messages.append({
                    "role": "DEBUG",
                    "content": debug_message,
                    "timestamp": turn.get("startTime", "")
                })
        elif turn["type"] == "request":
            if "payload" in turn and "query" in turn["payload"].get("payload", {}):
                messages.append({
                    "role": "HUMAN",
                    "content": turn["payload"]["payload"]["query"],
                    "timestamp": turn.get("startTime", "")
                })
        elif turn["type"] == "text" and "payload" in turn and "message" in turn["payload"].get("payload", {}):
            messages.append({
                "role": "BOT",
                "content": turn["payload"]["payload"]["message"],
                "timestamp": turn.get("startTime", "")
            })
    return messages

def save_transcript_to_txt(transcript_id, messages):
    filename = os.path.join(OUTPUT_DIRECTORY, f"transcript_{transcript_id}.txt")
    with open(filename, 'w', encoding='utf-8') as file:
        for message in messages:
            if message['role'] == 'DEBUG':
                file.write(f"DEBUG: {message['content']}\n")
            else:
                file.write(f"{message['role']}: {message['content']}\n")
            file.write("----------\n")  # Přidá oddělovací čáru
    print(f"Transkript uložen do {filename}")

def count_human_occurrences():
    human_count = 0
    for filename in os.listdir(OUTPUT_DIRECTORY):
        if filename.startswith("transcript_") and filename.endswith(".txt"):
            file_path = os.path.join(OUTPUT_DIRECTORY, filename)
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                human_count += content.count("HUMAN:")
    return human_count

def count_category_occurrences():
    category_counts = {category: 0 for category in CATEGORIES}
    
    for filename in os.listdir(OUTPUT_DIRECTORY):
        if filename.startswith("transcript_") and filename.endswith(".txt"):
            file_path = os.path.join(OUTPUT_DIRECTORY, filename)
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                for category in CATEGORIES:
                    exact_match = f'\\"{category}\\"'
                    category_counts[category] += content.count(exact_match)
    
    return category_counts

def create_pie_chart(category_counts):
    labels = list(category_counts.keys())
    sizes = list(category_counts.values())
    
    plt.figure(figsize=(10, 8))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    plt.axis('equal')
    plt.title('Rozložení kategorií')
    
    # Uložení grafu jako obrázek
    img_path = os.path.join(OUTPUT_DIRECTORY, 'category_distribution.png')
    plt.savefig(img_path)
    plt.close()
    
    return img_path

def create_custom_donut_chart(category_counts):
    # Seřazení kategorií sestupně podle počtu
    sorted_categories = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)
    labels = [cat for cat, count in sorted_categories if count > 0]
    sizes = [count for _, count in sorted_categories if count > 0]

    # Vytvoření širšího spektra barev duhy v opačném pořadí
    num_colors = len(labels)
    rainbow_colors = plt.cm.rainbow(np.linspace(1, 0, num_colors))

    # Změna velikosti figury na čtvercový tvar
    fig, ax = plt.subplots(figsize=(20, 20), subplot_kw=dict(aspect="equal"))  

    wedges, texts, autotexts = ax.pie(sizes, wedgeprops=dict(width=0.5), startangle=-40,
                                      colors=rainbow_colors, autopct='%1.1f%%', pctdistance=0.85)

    bbox_props = dict(boxstyle="round,pad=0.3", fc="w", ec="k", lw=0.72)
    kw = dict(arrowprops=dict(arrowstyle="-", connectionstyle="angle,angleA=0,angleB=90,rad=10"),
              bbox=bbox_props, zorder=0, va="center")

    # Úprava umístění popisků
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1) / 2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = f"angle,angleA=0,angleB={ang}"
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        ax.annotate(labels[i], xy=(x, y), xytext=(2.2*np.sign(x), 2.2*y),
                    horizontalalignment=horizontalalignment, fontsize=14, **kw)

    plt.title(f"Rozložení kategorií\n{START_DATE} - {END_DATE}", fontsize=24, y=1.05)

    for autotext in autotexts:
        autotext.set_visible(False)

    # Odstranění os pro čistší vzhled
    ax.set_axis_off()

    # Uložení grafu
    img_path = os.path.join(OUTPUT_DIRECTORY, 'custom_category_distribution.png')
    plt.savefig(img_path, bbox_inches='tight', dpi=300)
    plt.close()

    return img_path

def create_excel_report(ai_responses_count, category_counts):
    # Vytvoření nového Excel sešitu
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Nastavení nadpisu
    ws['A1'] = f"Report kategorií za období {START_DATE} - {END_DATE}"
    ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)

    # Přidání počtu AI odpovědí
    ws['A2'] = "Počet AI odpovědí utracených za dané období:"
    ws['B2'] = ai_responses_count

    # Zvýraznění počtu AI odpovědí
    ws['A2'].font = openpyxl.styles.Font(size=14, bold=True, color="FF0000")
    ws['B2'].font = openpyxl.styles.Font(size=14, bold=True, color="FF0000")
    ws['A2'].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws['B2'].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Přidání informace o celkovém počtu kategorizací
    total_categorizations = sum(category_counts.values())
    ws['A4'] = f"Celkový počet přiřazení / kategorizací dotazů: {total_categorizations}"
    ws['A5'] = "Poznámka: 1 dotaz může být přiřazen do více kategorií, proto se celkový počet AI odpovědí nerovná počtu kategorizací."

    # Přidání hlavičky tabulky
    ws['A7'] = "Kategorie"
    ws['B7'] = "Počet"
    for cell in ws['A7:B7'][0]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Seřazení kategorií sestupně podle počtu
    sorted_categories = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)

    # Přidání dat kategorií
    for row, (category, count) in enumerate(sorted_categories, start=8):
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=count)

    # Ohraničení tabulky
    max_row = ws.max_row
    for row in ws[f'A7:B{max_row}']:
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))

    # Nastavení šířky prvního sloupce
    ws.column_dimensions['A'].width = 150

    # Vytvoření vlastního grafu
    img_path = create_custom_donut_chart(category_counts)

    # Vložení obrázku do Excel souboru pod tabulkou
    img = Image(img_path)
    img.width = 1000
    img.height = 800
    ws.add_image(img, f'A{max_row + 2}')

    # Uložení Excel souboru
    output_filename = f'report_{START_DATE}_to_{END_DATE}.xlsx'
    output_path = os.path.join(OUTPUT_DIRECTORY, output_filename)
    wb.save(output_path)

    print(f"Nový Excel report byl vytvořen a uložen: {output_path}")

def print_summary(human_count, category_counts):
    print("\nSouhrn zpracování:")
    print(f"1) Počet HUMAN výskytů: {human_count}")
    print("\n2) Počty výskytů kategorií:")
    for category, count in category_counts.items():
        print(f"   {category}: {count}")

def main():
    create_output_directory()
    transcript_ids = get_transcript_ids()
    
    total_human_count = 0
    
    for transcript_id in transcript_ids:
        dialog = get_transcript_dialog(transcript_id)
        messages = extract_messages(dialog)
        save_transcript_to_txt(transcript_id, messages)
        
        for message in messages:
            if message['role'] == 'HUMAN':
                total_human_count += 1
    
    category_counts = count_category_occurrences()
    create_excel_report(total_human_count, category_counts)
    
    print("Zpracování dokončeno.")

if __name__ == "__main__":
    main()